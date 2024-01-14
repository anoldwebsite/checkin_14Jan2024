# https://www.geeksforgeeks.org/python-reading-excel-file-using-openpyxl-module/
import datetime
import openpyxl
from openpyxl import Workbook
import email_it
import pandas as pd


def make_excel_file(file_name="data_to_import.xlsx"):
    wb = Workbook()  # Create a workbook
    ws = wb.active  # Get the active worksheet
    header_row = tuple(['Artikel', 'Bezeichnung', 'Menge', 'Seriennummer', 'Ziellager', 'Lagerort', 'Preis'])
    ws.append(header_row)
    wb.save(file_name)


def add_row(r, file_name="data_to_import.xlsx"):
    """
    Adds a row to the Excel file.
    :param r: The row to add to the Excel file.
    :param file_name: The Excel file which would be opened.
    :return:
    """
    try:
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active
        r = tuple(r)
        ws.append(r)
        wb.save(file_name)
    except FileNotFoundError:
        print("Wrong File Path or File does not exist!")


def add_rows(rows_list, file_name):
    """
    Adds rows to the Excel file.
    :param rows_list:
    :param file_name: The Excel file which would be opened.
    :return:
    """
    try:
        wb = openpyxl.load_workbook(file_name)
        ws = wb.active
        if rows_list:
            for row in rows_list:
                r = tuple(row)
                ws.append(r)

            wb.save(file_name)
        else:
            print("No rows were added to the Excel File because an empty list was passed to function add_rows.")
    except FileNotFoundError:
        print("Wrong File Path or File does not exist!")


# Read Excel file
def read_excel_file(f):
    wb = openpyxl.load_workbook(f)
    ws = wb.active
    max_col = ws.max_column
    max_row = ws.max_row
    # Will print a particular row value
    for i in range(1, max_row + 1):
        for j in range(1, max_col + 1):
            ws.cell(row=i, column=j)
            # print(cell_obj.value)


def make_checkin_excel_file(machines):
    now = datetime.datetime.now()
    prefix = "data_to_import_{}-{}-{}-{}{}{}".format(now.year, now.month, now.day, now.hour, now.minute, now.second)
    file_name = "{}.xlsx".format(prefix)
    make_excel_file(file_name)
    add_rows(machines, file_name)
    # read_excel_file(file_name)
    email_it.send_excel_file(file_name)


def make_sn_substate_excel_file(sn_substate_dict):
    now = datetime.datetime.now()
    prefix = "data_to_import_{}-{}-{}-{}{}{}".format(now.year, now.month, now.day, now.hour, now.minute, now.second)
    file_name = "{}.xlsx".format(prefix)

    df = pd.DataFrame(list(sn_substate_dict.items()), columns=["Serial number", "Substate"])

    wb = Workbook()
    ws = wb.active

    # Write header
    header_row = ["Serial number", "Substate"]
    ws.append(header_row)

    # Write data
    for sn, substate in sn_substate_dict.items():
        ws.append([sn, substate])

    wb.save(file_name)

    print(f"Excel file '{file_name}' created with both serial numbers and substates.")


def make_sn_excel_file(sn_list):
    now = datetime.datetime.now()
    prefix = "data_to_import_{}-{}-{}-{}{}{}".format(now.year, now.month, now.day, now.hour, now.minute, now.second)
    file_name = "{}.xlsx".format(prefix)
    wb = Workbook()  # Create a workbook
    ws = wb.active  # Get the active worksheet
    header_row = tuple(['Artikel'])
    ws.append(header_row)
    if sn_list:
        for sn in sn_list:
            ws.append([sn])
    else:
        print("No rows were added to the Excel File because an empty list was passed to function make_sn_excel_file.")
    wb.save(file_name)
